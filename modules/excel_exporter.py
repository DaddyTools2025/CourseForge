import openpyxl
import re

class ExcelExporter:
    # 匹配选项开头的 "A. ", "B、", "C:", 等等
    _OPTION_PREFIX_RE = re.compile(r'^[A-Z][\.\、\:\s]\s*')

    @staticmethod
    def export_lms_questions(template_path: str, output_path: str, questions: list, course_name: str = "微课模块"):
        """
        无损装填 LMS 题库。禁用 pandas。使用 openpyxl 垂直多行写入。
        
        列映射关系:
        A=课程名称, B=章节名称, C=知识点名称, D=序号, E=题干及答案, F=正确答案(是/否), G=题型
        
        【垂直多行逻辑】:
        1题占 5 行 (1行题干 + 4行选项)
        第1行: A/B/C/D正常填, E填题干内容, G填题型(单选题)。F 留空。
        第2~5 行: E填选项文本, F填是否(是/否)。A,B,C,D,G 留空或沿用第一行数据。
        """
        if not questions:
            print("[ExcelExporter] 没有接收到任何题目，跳过导出")
            return

        try:
            wb = openpyxl.load_workbook(template_path)
            # 假设默认读活跃的表单
            ws = wb.active
        except Exception as e:
            raise FileNotFoundError(f"加载 Excel 模板失败，请确认导入模板文件是否存在: {e}")
            
        start_row = 2  # 假设第一行是表头
        
        for i, q in enumerate(questions):
            q_type = q.get("type", "单选题")
            
            # ====== 第 1 行: 题干行 ======
            # 优先使用题目内嵌的元数据，回退到函数参数
            q_course = q.get("course_name", "") or course_name
            q_chapter = q.get("chapter", "")
            q_section = q.get("section", "")
            
            ws.cell(row=start_row, column=1, value=q_course)                                # A: 课程名称
            ws.cell(row=start_row, column=2, value=q_chapter)                                # B: 章节名称
            ws.cell(row=start_row, column=3, value=q_section)                                # C: 知识点名称
            ws.cell(row=start_row, column=4, value=i + 1)                                  # D: 序号
            ws.cell(row=start_row, column=5, value=q.get("question", ""))                  # E: 题干
            ws.cell(row=start_row, column=7, value=q_type)                                 # G: 题型
            
            # 根据题型处理
            if q_type in ["判断题", "填空题"]:
                ws.cell(row=start_row, column=6, value=q.get("answer_content", ""))        # F: 答案
                start_row += 1
            else:
                ws.cell(row=start_row, column=6, value="")                                 # F: 题干行对应答案留空
                # ====== 第 2~5 行: 选项行 ======
                options = q.get("options", [])
                for j, opt in enumerate(options):
                    opt_row = start_row + 1 + j
                    
                    # E 列写选项内容，去除开头的 "A. " 等前缀，防止与LMS重复
                    raw_text = opt.get("text", "")
                    clean_text = ExcelExporter._OPTION_PREFIX_RE.sub('', raw_text)
                    ws.cell(row=opt_row, column=5, value=clean_text)              # E: 选项文本
                    
                    is_correct_val = opt.get("is_correct", False)
                    # 容错处理：如果是字符串格式
                    if isinstance(is_correct_val, str):
                        is_correct_bool = is_correct_val.lower() in ['true', '1', '是', 'yes']
                    else:
                        is_correct_bool = bool(is_correct_val)
                        
                    ws.cell(row=opt_row, column=6, value="是" if is_correct_bool else "")  # F: 若错误留空
                    
                # 游标下移：题干本身占1行 + 选项个数占N行
                start_row += (1 + len(options))
            
        try:
            wb.save(output_path)
        except PermissionError:
            raise PermissionError("文件被其他程序占用，请在 Excel 中关闭目标文件后重试！")
        except Exception as e:
            raise Exception(f"保存导出 Excel 时发生未知错误: {e}")
